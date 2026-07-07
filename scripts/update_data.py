import os
import requests
import subprocess
import time
from datetime import datetime, timezone, timedelta

# ====================== 설정 ======================
os.chdir(os.getenv('GITHUB_WORKSPACE', '.'))

DATA_DIR  = "scripts"
DATA_FILE = os.path.join(DATA_DIR, "data.xlsx")
TEMP_FILE = os.path.join(DATA_DIR, "temp_new.xlsx")

EMAIL    = os.getenv('MYHUBON_EMAIL')
PASSWORD = os.getenv('MYHUBON_PASSWORD')

if not EMAIL or not PASSWORD:
    print("❌ MYHUBON_EMAIL 또는 MYHUBON_PASSWORD secret이 설정되지 않았습니다.")
    exit(1)

# ====================== 공통 헤더 ======================
BASE_HEADERS = {
    "Content-Type":       "application/json",
    "User-Agent":         "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0",
    "sec-ch-ua":          '"Chromium";v="146", "Not-A.Brand";v="24", "Microsoft Edge";v="146"',
    "sec-ch-ua-mobile":   "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Origin":             "https://seller.myhubon.com",
    "Referer":            "https://seller.myhubon.com/",
    "X-Client-Type":      "ADMIN",
    "accept":             "application/json, text/plain, */*",
}

# ====================== STEP 1: 로그인 ======================
print("🔐 로그인 중...")

login_resp = requests.post(
    "https://api.myhubon.com/auth/login",
    headers=BASE_HEADERS,
    json={"email_id": EMAIL, "password": PASSWORD},
    timeout=30
)

if login_resp.status_code != 200:
    print(f"❌ 로그인 실패: {login_resp.status_code}")
    print("응답:", login_resp.text[:800])
    exit(1)

login_data = login_resp.json()

def extract_token(data):
    candidates = [
        data.get("state", {}).get("accessToken")  if isinstance(data.get("state"), dict) else None,
        data.get("state", {}).get("access_token") if isinstance(data.get("state"), dict) else None,
        data.get("accessToken"),
        data.get("access_token"),
        data.get("token"),
        data.get("data", {}).get("accessToken")   if isinstance(data.get("data"), dict) else None,
        data.get("data", {}).get("access_token")  if isinstance(data.get("data"), dict) else None,
    ]
    return next((t for t in candidates if t), None)

token = extract_token(login_data)

if not token:
    print("❌ 응답에서 토큰을 찾을 수 없습니다.")
    print("전체 응답 JSON:", login_resp.text[:1000])
    exit(1)

print(f"✅ 로그인 성공 (토큰 앞 20자: {token[:20]}...)")

# ====================== STEP 2: 최근 7일 ISO 형식으로 엑셀 다운로드 ======================
now_utc = datetime.now(timezone.utc)
to_dt   = now_utc.replace(hour=14, minute=59, second=59, microsecond=999000)
from_dt = (now_utc - timedelta(days=7)).replace(hour=15, minute=0, second=0, microsecond=0)

from_date_iso = from_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
to_date_iso   = to_dt.strftime("%Y-%m-%dT%H:%M:%S.999Z")

print(f"📅 조회 기간: {from_date_iso}  ~  {to_date_iso}")
print("📥 최근 7일 엑셀 다운로드 시작...")

download_headers = {**BASE_HEADERS, "Authorization": f"Bearer {token}"}
payload = {
    "from_date": from_date_iso,
    "to_date":   to_date_iso,
}

os.makedirs(DATA_DIR, exist_ok=True)

dl_resp = None
for attempt in range(1, 4):
    try:
        print(f"📡 시도 {attempt}/3...")
        dl_resp = requests.post(
            "https://api.myhubon.com/admin/product/manage/excel-download",
            headers=download_headers,
            json=payload,
            timeout=120
        )
        if dl_resp.status_code == 200:
            break
        print(f"⚠️ {dl_resp.status_code} 응답 → {'재시도' if attempt < 3 else '최종 실패'}")
        print("응답 앞 500자:", dl_resp.text[:500])
    except requests.exceptions.Timeout:
        print(f"⏱️ Timeout → {'재시도' if attempt < 3 else '최종 실패'}")

    if attempt < 3:
        wait = 30 * attempt
        print(f"⏳ {wait}초 대기...")
        time.sleep(wait)

if dl_resp is None or dl_resp.status_code != 200:
    print("❌ 다운로드 최종 실패")
    if dl_resp:
        print("응답:", dl_resp.text[:500])
    exit(1)

# ====================== STEP 3: 임시 파일에 저장 ======================
with open(TEMP_FILE, "wb") as f:
    f.write(dl_resp.content)

new_size_bytes = os.path.getsize(TEMP_FILE)
new_size_kb = new_size_bytes / 1024
print(f"✅ 신규 데이터 수신 완료 ({new_size_kb:.1f} KB)")

# ====================== STEP 4: 기존 엑셀과 병합 (Upsert) ======================
try:
    import openpyxl
except ImportError:
    print("📦 openpyxl 설치 중...")
    import subprocess as sp
    sp.run(["pip", "install", "openpyxl", "--break-system-packages", "-q"], check=True)
    import openpyxl

def merge_excel(existing_path, new_path, output_path):
    """기존 엑셀에 새 데이터를 Upsert 병합. 키: 첫 번째 열 값."""
    # [수정] 다운로드 파일 크기가 0이거나 비정상적인 포맷(Zip에러)일 때의 예외 처리 추가
    if not os.path.exists(new_path) or os.path.getsize(new_path) == 0:
        print("⚠️ 신규 다운로드 파일이 비어 있습니다. (0바이트)")
        return False

    try:
        wb_new = openpyxl.load_workbook(new_path)
    except Exception as e:
        print(f"❌ 신규 엑셀 파일을 읽을 수 없습니다. (포맷 에러: {e})")
        # 다운로드된 내용이 텍스트(에러 메시지 등)일 수 있으므로 로그 출력
        try:
            with open(new_path, "r", encoding="utf-8", errors="ignore") as f:
                print("📝 받은 파일 내용 앞부분:", f.read(300))
        except:
            pass
        return False

    ws_new = wb_new.active
    new_rows = list(ws_new.iter_rows(values_only=True))
    if not new_rows or len(new_rows) <= 1:
        print("⚠️ 신규 데이터 시트가 비어 있거나 헤더만 존재합니다. 병합 스킵.")
        return False

    header = new_rows[0]
    data_rows = new_rows[1:]
    print(f"  신규 데이터: 헤더 {len(header)}열, 데이터 {len(data_rows)}행")

    if not os.path.exists(existing_path):
        import shutil
        shutil.copy2(new_path, output_path)
        print("  기존 파일 없음 → 신규 데이터 그대로 사용")
        return True

    wb_ex = openpyxl.load_workbook(existing_path)
    ws_ex = wb_ex.active
    ex_rows = list(ws_ex.iter_rows(values_only=True))
    ex_header = ex_rows[0] if ex_rows else None

    if ex_header != header:
        print(f"  ⚠️ 헤더 불일치 (기존: {str(ex_header)[:60]} / 신규: {str(header)[:60]})")
        print("  → 기존 파일을 신규 파일로 교체합니다.")
        import shutil
        shutil.copy2(new_path, output_path)
        return True

    # 기존 행을 키(첫 번째 열) 기준으로 dict에 로드
    ex_data = {}
    for row in ex_rows[1:]:
        key = row[0]
        if key is not None:
            ex_data[key] = row

    # 신규 데이터 upsert
    added = updated = 0
    for row in data_rows:
        key = row[0]
        if key is None:
            continue
        if key in ex_data:
            if ex_data[key] != row:
                ex_data[key] = row
                updated += 1
        else:
            ex_data[key] = row
            added += 1

    print(f"  병합 결과: 신규 추가 {added}행 / 업데이트 {updated}행 / 합계 {len(ex_data)}행")

    # 결과 워크북 생성
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(list(header))
    try:
        sorted_keys = sorted(ex_data.keys(), key=lambda k: (str(type(k).__name__), str(k)))
    except TypeError:
        sorted_keys = list(ex_data.keys())
    for key in sorted_keys:
        ws_out.append(list(ex_data[key]))

    wb_out.save(output_path)
    return True


MERGED_FILE = os.path.join(DATA_DIR, "merged.xlsx")

# [수정] 수신된 데이터가 아예 없거나(0바이트) 병합 작업이 필요 없는 경우에 대한 흐름 제어
if new_size_bytes == 0:
    print("ℹ️ 신규 데이터가 없어 병합 및 업데이트를 진행하지 않습니다.")
    if os.path.exists(TEMP_FILE):
        os.remove(TEMP_FILE)
    success = False
    # 에러 종료가 아닌 정상 종료를 원하므로 exit(0) 처리
    exit(0)
else:
    success = merge_excel(DATA_FILE, TEMP_FILE, MERGED_FILE)

if not success:
    print("❌ 병합 실패 또는 스킵 → 기존 파일 유지, 정상 종료 처리")
    if os.path.exists(TEMP_FILE):
        os.remove(TEMP_FILE)
    # 병합이 실패하거나 대상이 아닐 때 Actions 워크플로우 전체가 깨지지 않도록 exit(0)로 처리합니다.
    exit(0)

# 병합 파일로 교체
if os.path.exists(DATA_FILE):
    os.remove(DATA_FILE)
os.rename(MERGED_FILE, DATA_FILE)
if os.path.exists(TEMP_FILE):
    os.remove(TEMP_FILE)

final_size = os.path.getsize(DATA_FILE) / (1024 * 1024)
print(f"✅ {DATA_FILE} 저장 완료 ({final_size:.2f} MB)")

# ====================== STEP 5: Git Push ======================
today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

subprocess.run(["git", "config", "user.name",  "Data Updater Bot"],       check=True)
subprocess.run(["git", "config", "user.email", "bot@jjangse1.github.io"], check=True)
subprocess.run(["git", "add", DATA_FILE], check=True)

result = subprocess.run(["git", "diff", "--cached", "--quiet"], capture_output=True)

if result.returncode != 0:
    subprocess.run(["git", "commit", "-m", f"auto: data.xlsx 업데이트 ({today})"], check=True)
    subprocess.run(["git", "push"], check=True)
    print("✅ Git commit & push 완료")
else:
    print("ℹ️ 변경사항 없음 → push 생략")
